# Charles Shaviro
# Center for an Urban Future
# Created 10/1/18
# Edited 5/14/20

# CUF Research 
# Reverse Geocoding Task
# 
# This program processes our data on NYC firms, leveraging  
# Mapbox's Python SDK to look up the neighborhood each firm is located in 
# through 'reverse geocoding' (using lat/long data to pull a text description
# of information such as the address, country, and neighborhood).
#
#



# importing openpyxl library that allows us to read/write from/to excel files
from openpyxl import Workbook
from openpyxl import load_workbook

# importing mapbox library to let us reverse geocode
from mapbox import Geocoder

# importing Numpy python module for its bevy of statistical functions
import numpy


class Dataset():
    '''
    Worksheet must have columns going out to 'EK', if not fewer.
    NOTE: If the excel uses more columns beyond column 'EK', they will not be 
    included, -> we will need to change the declaration of the index variable.
    The instance variable containing the actual data has been named 'data'.
    '''
    
    def __init__(self, filename):
        
        '''
        In initiating a Dataset class, we store the desired worksheet in an 
        instance variable called 'self.data'.
        '''
        
        workbook = load_workbook(filename)
        worksheet = workbook['Sheet1']
        
        # declaration of 'index' variable- it is the farthest column that we 
        # want to include (more columns --> slower computations)
        index = 'EK'+str(worksheet.max_row)
        
        # indexing the desired columns from the sheet, we put them in an 
        # instance variable
        self.data = worksheet['A1':index]


def write_to_excel(data_matrix, filename):
    """
        Writes a 2-d matrix to an excel spreadsheet.
    """
    wb = Workbook()
    ws = wb.create_sheet("data", 0)
    
    for row in data_matrix:
        ws.append(row)
    
    save_file = filename + '.xlsx'
    wb.save( save_file )



def reverse_geocode_all_neighborhoods( firms ):
    '''
    This function loops through our NYC firm data, which must be contained in a 'Database'
    when passed as an argument. The function relies upon the library provided by 'Mapbox' 
    to use each firm's latitude and longitude to look up geographical data, and return
    the neighborhood of each coordinate point/business. These data, along with each 
    startup's original industry, latitude, and longitude, are outputted.
    '''
    
    
    # first we set up our geocoder
    geocoder = Geocoder(access_token="Paste your access code in here")
    
    output = [ ["Industry", "Latitude", "Longitude", "Neighborhood"] ]
    
    
    # establish the progress variable, for testing
    progress = 0
    
    for row in firms.data[1:]:
        
        # as there are 854 entries in our firm data, this will print 'processing' 20 times, 
        # at each ~5% progress point of our function
        progress += 1
        if (progress % 43 == 0):
            print "processing"        
        
        output_row = []
        
        # the industry, latidude, and longitude are contained in the
        # three cells of each row
        output_row.append( row[0].value )
        output_row.append( row[1].value )
        output_row.append( row[2].value )
        
        
        local_lat = row[1].value
        local_long = row[2].value


        # passing the lat/long to the geocoder, we pull our geo-data
        response = geocoder.reverse(lon=local_long, lat=local_lat)
        initial_collection = response.json()
        
        features = initial_collection['features']
        
        # we get our neighborhood text from the json of geo-data
        neighborhood = str(features[0]['context'][0]['text'])
        
        # we add the neighborhood to this row of output
        output_row.append( neighborhood )
        
        # with all info included, we append our row to the output matrix
        output.append( output_row )
    
    return output
        
        